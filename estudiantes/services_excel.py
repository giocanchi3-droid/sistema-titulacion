import logging
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl import Workbook

from .models import HistorialExpediente, Programa, RegistroTitulacion


logger = logging.getLogger(__name__)


EXPORTAR_CAMPOS = [
    ("ID BANNER", "id_banner"),
    ("NOMBRES COMPLETOS", "nombres_completos"),
    ("CEDULA", "cedula"),
    ("CELULAR", "celular"),
    ("CORREO PERSONAL", "correo_personal"),
    ("CORREO INSTITUC", "correo_instituc"),
    ("SEDE", "sede"),
    ("PROGRAMA", "programa"),
    ("PROGAMA DESC", "programa_desc"),
    ("NUMERO DE COHORTE", "numero_cohorte"),
    ("PERIODO DE INGRESO", "periodo_ingreso"),
    ("NIVEL2", "nivel2"),
    ("MODALIDAD DE TITULACION", "modalidad_titulacion"),
    ("MATRICULA UIC", "matricula_uic"),
    ("PERIODO DE TITULACION SENESCYT", "periodo_titulacion_senescyt"),
    ("ESTADO", "estado"),
    ("CUMPLIMIENTO DE IDIOMA", "cumplimiento_idioma"),
    ("MATERIA PRACTICAS PRE PROFESIONALES", "materia_practicas_pre_profesionales"),
    ("HORAS 240", "horas_240"),
    ("MATERIA SERVICIO COMUNITARIO", "materia_servicio_comunitario"),
    ("HORAS 120", "horas_120"),
    ("NOMBRES COMPLETOS TUTOR", "nombres_completos_tutor"),
    ("ID TUTOR", "id_tutor"),
    ("TEMA", "tema"),
    ("1ER MIEMBREO DE TRIBUNAL APELLIDOS Y NOMBRES COMPLETOS", "primer_miembro_tribunal"),
    ("1ER MIEMBRO DE TRIBUNAL ID DOCENTE", "primer_miembro_id_docente"),
    ("2DO MIEMBRO DE TRIBUNAL APELLIDOS Y NOMBRES COMPLETOS", "segundo_miembro_tribunal"),
    ("2DO MIEMBRO DE TRIBUNAL ID DOCENTE", "segundo_miembro_id_docente"),
    ("3TER MIEMBRO DE TRIBUNAL NOMBRES COMPLETOS", "tercer_miembro_tribunal"),
    ("3TER MIEMBRO DE TRIBUNAL ID DOCENTE", "tercer_miembro_id_docente"),
    ("4TO MIEMBRO DE TRIBUNAL", "cuarto_miembro_tribunal"),
    ("4TO MIEMBRO DE TRIBUNAL ID DOCENTE", "cuarto_miembro_id_docente"),
    ("PROYECTO ESCRITO", "proyecto_escrito"),
    ("DEFENSA ORAL", "defensa_oral"),
    ("NOTA FINAL", "nota_final"),
    ("EXAMEN TEORICO COMPLEXIVO", "examen_teorico_complexivo"),
    ("EXAMEN TEORICO PRACTICO", "examen_teorico_practico"),
    ("NOTA FINAL2", "nota_final2"),
    ("OBSERVACION PUCE TEC", "observacion_puce_tec"),
    ("OBSERVACIONES DE SECRETARIA GENERAL", "observaciones_secretaria_general"),
    ("NUEVA OBSERVACION PUCE TEC", "nueva_observacion_puce_tec"),
    ("ESTADO DE ENVIO DE REGISTRO", "estado_envio_registro"),
    ("FECHA DE GRADO", "fecha_grado"),
    ("OBSERVACION SECRETARIA", "observacion_secretaria"),
]


ENCABEZADOS_ESPERADOS = {
    "ID BANNER",
    "NOMBRES COMPLETOS",
    "CEDULA",
    "CELULAR",
    "CORREO PERSONAL",
    "CORREO INSTITUC",
    "SEDE",
    "PROGRAMA",
    "PROGAMA DESC",
    "NUMERO DE COHORTE",
    "PERIODO DE INGRESO",
    "NIVEL2",
    "MODALIDAD DE TITULACION",
    "MATRICULA UIC",
    "PERIODO DE TITULACION SENESCYT",
    "ESTADO",
    "CUMPLIMIENTO DE IDIOMA",
    "MATERIA PRACTICAS PRE PROFESIONALES",
    "HORAS 240",
    "MATERIA SERVICIO COMUNITARIO",
    "HORAS 120",
    "NOMBRES COMPLETOS TUTOR",
    "ID TUTOR",
    "TEMA",
    (
        "1ER MIEMBREO DE TRIBUNAL "
        "APELLIDOS Y NOMBRES COMPLETOS"
    ),
    "1ER MIEMBRO DE TRIBUNAL ID DOCENTE",
    (
        "2DO MIEMBRO DE TRIBUNAL "
        "APELLIDOS Y NOMBRES COMPLETOS"
    ),
    "2DO MIEMBRO DE TRIBUNAL ID DOCENTE",
    (
        "3TER MIEMBRO DE TRIBUNAL "
        "NOMBRES COMPLETOS"
    ),
    "3TER MIEMBRO DE TRIBUNAL ID DOCENTE",
    "4TO MIEMBRO DE TRIBUNAL",
    "4TO MIEMBRO DE TRIBUNAL ID DOCENTE",
    "PROYECTO ESCRITO",
    "DEFENSA ORAL",
    "NOTA FINAL",
    "EXAMEN TEORICO COMPLEXIVO",
    "EXAMEN TEORICO PRACTICO",
    "NOTA FINAL2",
    "OBSERVACION PUCE TEC",
    "OBSERVACIONES DE SECRETARIA GENERAL",
    "NUEVA OBSERVACION PUCE TEC",
    "ESTADO DE ENVIO DE REGISTRO",
    "FECHA DE GRADO",
    "OBSERVACION SECRETARIA",
}

COLUMN_ALIASES = {
    "id_banner": {
        "ID BANNER",
        "ID BANNER ",
        "ID_BANNER",
        "IDBANNER",
        "ID BANNER ESTUDIANTE",
    },
    "cedula": {
        "CEDULA",
        "CEDULA ESTUDIANTE",
        "CEDULA IDENTIDAD",
        "CÉDULA",
        "CI",
    },
    "nombres_completos": {
        "NOMBRES COMPLETOS",
        "NOMBRES",
        "NOMBRES COMPLETOS ESTUDIANTE",
        "NOMBRE COMPLETO",
        "NOMBRES COMPLETOS",
    },
    "celular": {"CELULAR", "TELEFONO", "TELÉFONO", "MOVIL"},
    "correo_personal": {"CORREO PERSONAL", "EMAIL PERSONAL", "CORREO ELECTRONICO PERSONAL"},
    "correo_instituc": {"CORREO INSTITUC", "CORREO INSTITUCIONAL", "EMAIL INSTITUCIONAL", "CORREO ELECTRONICO INSTITUCIONAL"},
    "sede": {"SEDE", "SEDE ACADEMICA", "SEDE PRINCIPAL"},
    "programa": {"PROGRAMA", "PROGRAMA"},
    "programa_desc": {"PROGAMA DESC", "PROGAMA_DESC", "PROGRAMA DESC", "DESCRIPCION PROGRAMA", "DESCRIPCIÓN PROGRAMA", "DESCRIPCION DEL PROGRAMA"},
    "numero_cohorte": {"NUMERO DE COHORTE", "NÚMERO DE COHORTE", "COHORTE"},
    "periodo_ingreso": {"PERIODO DE INGRESO", "PERIODO INGRESO"},
    "nivel2": {"NIVEL2", "NIVEL 2", "NIVEL"},
    "modalidad_titulacion": {"MODALIDAD DE TITULACION", "MODALIDAD DE TITULACIÓN", "MODALIDAD"},
    "matricula_uic": {"MATRICULA UIC", "MATRÍCULA UIC", "MATRICULA"},
    "periodo_titulacion_senescyt": {"PERIODO DE TITULACION SENESCYT", "PERIODO TITULACION SENESCYT", "PERIODO DE TITULACIÓN SENESCYT"},
    "estado": {"ESTADO", "ESTADO ESTUDIANTE"},
    "cumplimiento_idioma": {"CUMPLIMIENTO DE IDIOMA", "IDIOMA"},
    "materia_practicas_pre_profesionales": {"MATERIA PRACTICAS PRE PROFESIONALES", "MATERIA PRÁCTICAS PRE PROFESIONALES", "PRÁCTICAS PREPROFESIONALES"},
    "horas_240": {"HORAS 240", "H 240", "HORAS240"},
    "materia_servicio_comunitario": {"MATERIA SERVICIO COMUNITARIO", "MATERIA SERVICIO COMUNITARIO "},
    "horas_120": {"HORAS 120", "H 120", "HORAS120"},
    "nombres_completos_tutor": {"NOMBRES COMPLETOS TUTOR", "TUTOR"},
    "id_tutor": {"ID TUTOR", "ID_TUTOR"},
    "tema": {"TEMA", "TEMA PROYECTO"},
    "primer_miembro_tribunal": {"1ER MIEMBRO DE TRIBUNAL APELLIDOS Y NOMBRES COMPLETOS", "1ER MIEMBREO DE TRIBUNAL APELLIDOS Y NOMBRES COMPLETOS"},
    "primer_miembro_id_docente": {"1ER MIEMBRO DE TRIBUNAL ID DOCENTE", "ID DOCENTE 1"},
    "segundo_miembro_tribunal": {"2DO MIEMBRO DE TRIBUNAL APELLIDOS Y NOMBRES COMPLETOS", "2DO MIEMBREO DE TRIBUNAL APELLIDOS Y NOMBRES COMPLETOS"},
    "segundo_miembro_id_docente": {"2DO MIEMBRO DE TRIBUNAL ID DOCENTE", "ID DOCENTE 2"},
    "tercer_miembro_tribunal": {"3ER MIEMBRO DE TRIBUNAL NOMBRES COMPLETOS", "3TER MIEMBRO DE TRIBUNAL NOMBRES COMPLETOS"},
    "tercer_miembro_id_docente": {"3ER MIEMBRO DE TRIBUNAL ID DOCENTE", "3TER MIEMBRO DE TRIBUNAL ID DOCENTE", "ID DOCENTE 3"},
    "cuarto_miembro_tribunal": {"4TO MIEMBRO DE TRIBUNAL", "4TO MIEMBRO DE TRIBUNAL NOMBRES COMPLETOS"},
    "cuarto_miembro_id_docente": {"4TO MIEMBRO DE TRIBUNAL ID DOCENTE", "ID DOCENTE 4"},
    "proyecto_escrito": {"PROYECTO ESCRITO"},
    "defensa_oral": {"DEFENSA ORAL"},
    "nota_final": {"NOTA FINAL"},
    "examen_teorico_complexivo": {"EXAMEN TEORICO COMPLEXIVO", "EXAMEN TEÓRICO COMPLEXIVO"},
    "examen_teorico_practico": {"EXAMEN TEORICO PRACTICO", "EXAMEN TEÓRICO PRÁCTICO"},
    "nota_final2": {"NOTA FINAL2", "NOTA FINAL 2"},
    "observacion_puce_tec": {"OBSERVACION PUCE TEC", "OBSERVACIÓN PUCE TEC"},
    "observaciones_secretaria_general": {"OBSERVACIONES DE SECRETARIA GENERAL", "OBSERVACIONES DE SECRETARÍA GENERAL"},
    "nueva_observacion_puce_tec": {"NUEVA OBSERVACION PUCE TEC", "NUEVA OBSERVACIÓN PUCE TEC"},
    "estado_envio_registro": {"ESTADO DE ENVIO DE REGISTRO", "ESTADO DE ENVÍO DE REGISTRO"},
    "fecha_grado": {"FECHA DE GRADO", "FECHA GRADO"},
    "observacion_secretaria": {"OBSERVACION SECRETARIA", "OBSERVACIÓN SECRETARÍA", "OBSERVACION SECRETARIA GENERAL"},
}


def resolver_campo(encabezado):
    normalizado = normalizar(encabezado)

    for campo, aliases in COLUMN_ALIASES.items():
        if normalizado in aliases:
            return campo

    normalizado_sin_espacios = normalizado.replace(" ", "")
    for campo, aliases in COLUMN_ALIASES.items():
        alias_normalizado = {
            normalizar(alias)
            for alias in aliases
        }
        if normalizado_sin_espacios in {
            alias.replace(" ", "")
            for alias in alias_normalizado
        }:
            return campo

    return None


def normalizar(valor):
    """
    Normaliza encabezados eliminando acentos, saltos de línea,
    guiones, signos especiales y diferencias entre mayúsculas.
    """

    valor = str(valor or "").strip().upper()
    valor = unicodedata.normalize("NFD", valor)

    valor = "".join(
        caracter
        for caracter in valor
        if unicodedata.category(caracter) != "Mn"
    )

    valor = valor.replace("_", " ")
    valor = valor.replace("\n", " ")
    valor = valor.replace("\r", " ")

    valor = re.sub(
        r"[^A-Z0-9 ]",
        " ",
        valor,
    )

    valor = re.sub(
        r"\s+",
        " ",
        valor,
    )

    return valor.strip()


def convertir_texto(valor):
    if valor is None:
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()


def convertir_cedula(valor):
    if valor is None:
        return ""

    if isinstance(valor, (int, float)):
        cedula = str(int(valor))
        cedula = cedula.zfill(10)
    else:
        cedula = str(valor).strip()
        cedula = cedula.replace(".0", "")
        cedula = re.sub(r"\D", "", cedula)

    return cedula


def convertir_entero(valor, valor_predeterminado=None):
    if valor in (None, ""):
        return valor_predeterminado

    try:
        numero = int(float(valor))
    except (TypeError, ValueError) as error:
        raise ValueError(
            f"El valor '{valor}' no es un número entero."
        ) from error

    if numero < 0:
        raise ValueError(
            "Las horas no pueden ser negativas."
        )

    return numero


def convertir_nota(valor):
    if valor in (None, ""):
        return None

    try:
        nota = Decimal(
            str(valor).strip().replace(",", ".")
        )
    except (InvalidOperation, TypeError, ValueError) as error:
        raise ValueError(
            f"La nota '{valor}' no es válida."
        ) from error

    if nota < 0 or nota > 10:
        raise ValueError(
            f"La nota {nota} debe estar entre 0 y 10."
        )

    return nota


def convertir_fecha(valor):
    if valor in (None, ""):
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto_fecha = str(valor).strip()

    formatos = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(
                texto_fecha,
                formato,
            ).date()
        except ValueError:
            continue

    raise ValueError(
        f"La fecha '{texto_fecha}' no tiene un formato válido."
    )


def convertir_modalidad(valor):
    modalidad = normalizar(valor)

    if not modalidad:
        return ""

    if "COMPLEX" in modalidad:
        return "EXAMEN_COMPLEXIVO"

    if (
        "TRABAJO" in modalidad
        or "ESCRITO" in modalidad
        or "PROYECTO" in modalidad
    ):
        return "TRABAJO_ESCRITO"

    raise ValueError(
        f"La modalidad '{valor}' no está reconocida."
    )


def convertir_estado(valor):
    estado = normalizar(valor)

    equivalencias = {
        "": "",
        "REGISTRADO": "REGISTRADO",
        "REGISTRO": "REGISTRADO",
        "INSCRITO": "REGISTRADO",
        "INSCRIPCION": "REGISTRADO",
        "EN PROCESO": "EN_PROCESO",
        "PROCESO": "EN_PROCESO",
        "REVISION": "REVISION",
        "EN REVISION": "REVISION",
        "DEFENSA": "DEFENSA",
        "APROBADO": "APROBADO",
        "APROBADA": "APROBADO",
        "GRADUADO": "GRADUADO",
        "GRADUADA": "GRADUADO",
        "OBSERVADO": "OBSERVADO",
        "OBSERVADA": "OBSERVADO",
    }

    if estado not in equivalencias:
        raise ValueError(
            f"El estado '{valor}' no está reconocido."
        )

    return equivalencias[estado]


def convertir_cumplimiento(valor):
    cumplimiento = normalizar(valor)

    equivalencias = {
        "": "",
        "SI": "SI",
        "S": "SI",
        "CUMPLE": "SI",
        "APROBADO": "SI",
        "NO": "NO",
        "N": "NO",
        "NO CUMPLE": "NO",
        "PENDIENTE": "PENDIENTE",
        "EN PROCESO": "PENDIENTE",
    }

    if cumplimiento not in equivalencias:
        raise ValueError(
            f"El cumplimiento de idioma '{valor}' no es válido."
        )

    return equivalencias[cumplimiento]


def convertir_estado_envio(valor):
    estado = normalizar(valor)

    equivalencias = {
        "": "",
        "NO ENVIADO": "NO_ENVIADO",
        "PENDIENTE": "NO_ENVIADO",
        "ENVIADO": "ENVIADO",
        "OBSERVADO": "OBSERVADO",
        "OBSERVADA": "OBSERVADO",
        "APROBADO": "APROBADO",
        "APROBADA": "APROBADO",
    }

    if estado not in equivalencias:
        raise ValueError(
            f"El estado de envío '{valor}' no es válido."
        )

    return equivalencias[estado]


def obtener(datos, encabezado):
    """
    Obtiene un valor del diccionario utilizando:
    - la clave exacta,
    - el nombre normalizado,
    - el nombre interno del campo,
    - o cualquiera de sus aliases.
    """

    if not isinstance(datos, dict):
        return None

    if encabezado in datos:
        return datos[encabezado]

    clave_normalizada = normalizar(encabezado)

    for nombre, valor in datos.items():
        if normalizar(nombre) == clave_normalizada:
            return valor

    for campo, aliases in COLUMN_ALIASES.items():
        nombres_validos = {normalizar(campo)}
        nombres_validos.update(
            normalizar(alias)
            for alias in aliases
        )

        if clave_normalizada in nombres_validos:
            if campo in datos:
                return datos[campo]

    return None


def guardar_registro(datos, usuario=None):
    cedula = convertir_cedula(
        obtener(datos, "CEDULA")
    )
    id_banner = convertir_texto(
        obtener(datos, "ID_BANNER")
    )

    if not cedula and id_banner:
        registro_existente = (
            RegistroTitulacion.objects.filter(id_banner=id_banner)
            .order_by("-fecha_actualizacion")
            .first()
        )
        if registro_existente:
            cedula = registro_existente.cedula

    nombres = convertir_texto(
        obtener(datos, "NOMBRES COMPLETOS")
    )

    programa = convertir_texto(
        obtener(datos, "PROGRAMA")
    )

    if not cedula:
        raise ValueError(
            "La cédula está vacía."
        )

    if len(cedula) != 10 or not cedula.isdigit():
        raise ValueError(
            f"La cédula '{cedula}' debe tener 10 dígitos."
        )

    if not nombres:
        raise ValueError(
            "El nombre del estudiante está vacío."
        )

    if not programa:
        raise ValueError(
            "El programa está vacío."
        )

    valores = {
        "id_banner": convertir_texto(
            obtener(datos, "ID_BANNER")
        ),
        "nombres_completos": nombres,
        "celular": convertir_texto(
            obtener(datos, "CELULAR")
        ),
        "correo_personal": convertir_texto(
            obtener(datos, "CORREO_PERSONAL")
        ),
        "correo_instituc": convertir_texto(
            obtener(datos, "CORREO_INSTITUC")
        ),
        "sede": convertir_texto(
            obtener(datos, "SEDE")
        ),
        "programa": programa,
        "programa_desc": convertir_texto(
            obtener(datos, "PROGAMA_DESC")
        ),
        "numero_cohorte": convertir_texto(
            obtener(datos, "NÚMERO DE COHORTE")
        ),
        "periodo_ingreso": convertir_texto(
            obtener(datos, "PERIODO DE INGRESO")
        ),
        "nivel2": convertir_texto(
            obtener(datos, "NIVEL2")
        ),
        "modalidad_titulacion": convertir_modalidad(
            obtener(datos, "MODALIDAD DE TITULACIÓN")
        ),
        "matricula_uic": convertir_texto(
            obtener(datos, "MATRICULA UIC")
        ),
        "periodo_titulacion_senescyt": convertir_texto(
            obtener(
                datos,
                "PERIODO DE TITULACIÓN SENESCYT",
            )
        ),
        "estado": convertir_estado(
            obtener(datos, "ESTADO")
        ),
        "cumplimiento_idioma": convertir_cumplimiento(
            obtener(datos, "CUMPLIMIENTO DE IDIOMA")
        ),
        "materia_practicas_pre_profesionales": convertir_texto(
            obtener(
                datos,
                "MATERIA PRÁCTICAS PRE PROFESIONALES",
            )
        ),
        "horas_240": convertir_entero(
            obtener(datos, "HORAS 240"),
            240,
        ),
        "materia_servicio_comunitario": convertir_texto(
            obtener(
                datos,
                "MATERIA SERVICIO COMUNITARIO",
            )
        ),
        "horas_120": convertir_entero(
            obtener(datos, "HORAS 120"),
            120,
        ),
        "nombres_completos_tutor": convertir_texto(
            obtener(
                datos,
                "NOMBRES COMPLETOS TUTOR",
            )
        ),
        "id_tutor": convertir_texto(
            obtener(datos, "ID TUTOR")
        ),
        "tema": convertir_texto(
            obtener(datos, "TEMA")
        ),
        "primer_miembro_tribunal": convertir_texto(
            obtener(
                datos,
                (
                    "1er MIEMBREO DE TRIBUNAL "
                    "APELLIDOS Y NOMBRES COMPLETOS"
                ),
            )
        ),
        "primer_miembro_id_docente": convertir_texto(
            obtener(
                datos,
                "1er MIEMBRO DE TRIBUNAL ID DOCENTE",
            )
        ),
        "segundo_miembro_tribunal": convertir_texto(
            obtener(
                datos,
                (
                    "2do MIEMBRO DE TRIBUNAL "
                    "APELLIDOS Y NOMBRES COMPLETOS"
                ),
            )
        ),
        "segundo_miembro_id_docente": convertir_texto(
            obtener(
                datos,
                "2do MIEMBRO DE TRIBUNAL ID DOCENTE",
            )
        ),
        "tercer_miembro_tribunal": convertir_texto(
            obtener(
                datos,
                (
                    "3ter MIEMBRO DE TRIBUNAL "
                    "NOMBRES COMPLETOS"
                ),
            )
        ),
        "tercer_miembro_id_docente": convertir_texto(
            obtener(
                datos,
                "3ter MIEMBRO DE TRIBUNAL ID DOCENTE",
            )
        ),
        "cuarto_miembro_tribunal": convertir_texto(
            obtener(
                datos,
                "4to MIEMBRO DE TRIBUNAL",
            )
        ),
        "cuarto_miembro_id_docente": convertir_texto(
            obtener(
                datos,
                "4to MIEMBRO DE TRIBUNAL ID DOCENTE",
            )
        ),
        "proyecto_escrito": convertir_nota(
            obtener(datos, "PROYECTO ESCRITO")
        ),
        "defensa_oral": convertir_nota(
            obtener(datos, "DEFENSA ORAL")
        ),
        "nota_final": convertir_nota(
            obtener(datos, "NOTA FINAL")
        ),
        "examen_teorico_complexivo": convertir_nota(
            obtener(
                datos,
                "EXAMEN TEÓRICO COMPLEXIVO",
            )
        ),
        "examen_teorico_practico": convertir_nota(
            obtener(
                datos,
                "EXAMEN TEÓRICO PRÁCTICO",
            )
        ),
        "nota_final2": convertir_nota(
            obtener(datos, "NOTA FINAL2")
        ),
        "observacion_puce_tec": convertir_texto(
            obtener(datos, "OBSERVACIÓN PUCE TEC")
        ),
        "observaciones_secretaria_general": convertir_texto(
            obtener(
                datos,
                "OBSERVACIONES DE SECRETARÍA GENERAL",
            )
        ),
        "nueva_observacion_puce_tec": convertir_texto(
            obtener(
                datos,
                "NUEVA OBSERVACIÓN PUCE TEC",
            )
        ),
        "estado_envio_registro": convertir_estado_envio(
            obtener(
                datos,
                "ESTADO DE ENVÍO DE REGISTRO",
            )
        ),
        "fecha_grado": convertir_fecha(
            obtener(datos, "Fecha de Grado")
        ),
        "observacion_secretaria": convertir_texto(
            obtener(datos, "Observación Secretaría")
        ),
    }

    registro = RegistroTitulacion.objects.filter(
        cedula=cedula
    ).first()

    if registro is None and id_banner:
        registro = RegistroTitulacion.objects.filter(
            id_banner=id_banner
        ).order_by("-fecha_actualizacion").first()

    creado = registro is None

    anterior = {}

    if creado:
        registro = RegistroTitulacion(
            cedula=cedula,
            id_banner=id_banner,
        )
    else:
        anterior = {
            campo: str(getattr(registro, campo, "") or "")
            for campo in valores
        }

    if not registro.id_banner and id_banner:
        registro.id_banner = id_banner

    for campo, valor in valores.items():
        setattr(
            registro,
            campo,
            valor,
        )

    registro.full_clean()
    registro.save()

    if programa:
        descripcion_programa = (
            convertir_texto(valores.get("programa_desc"))
            or (Programa.objects.filter(codigo=programa).first().descripcion if Programa.objects.filter(codigo=programa).exists() else programa)
        )
        Programa.objects.update_or_create(
            codigo=programa,
            defaults={
                "descripcion": descripcion_programa,
                "activo": True,
            },
        )

    responsable = getattr(usuario, "username", None) or "Importación Excel"
    for campo, valor in valores.items():
        nuevo = str(valor or "")
        if creado or anterior.get(campo, "") != nuevo:
            HistorialExpediente.objects.create(
                registro=registro,
                registro_nombre=registro.nombres_completos,
                registro_cedula=registro.cedula,
                responsable=responsable,
                campo=registro._meta.get_field(campo).verbose_name,
                valor_anterior=anterior.get(campo, ""),
                valor_nuevo=nuevo,
                accion="IMPORTACION" if creado else "EDICION",
            )

    return registro, creado


def importar_excel(archivo, usuario=None):
    try:
        libro = load_workbook(
            archivo,
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise ValueError(
            "El archivo no pudo ser leído como una matriz Excel."
        ) from error

    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))

    if not filas or not any(
        valor not in (None, "")
        for valor in filas[0]
    ):
        libro.close()
        raise ValueError(
            "El archivo Excel está vacío."
        )

    encabezados_originales = filas[0]
    encabezados = [
        normalizar(encabezado)
        for encabezado in encabezados_originales
    ]

    campos_encontrados = set()
    for encabezado in encabezados:
        campo = resolver_campo(encabezado)
        if campo:
            campos_encontrados.add(campo)

    if not campos_encontrados:
        libro.close()
        raise ValueError(
            "La matriz no contiene encabezados reconocidos para importar estudiantes."
        )

    identificadores_requeridos = {"cedula", "id_banner"}
    if not identificadores_requeridos.intersection(campos_encontrados):
        libro.close()
        raise ValueError(
            "La matriz debe incluir al menos una columna de identificación (CÉDULA o ID_BANNER)."
        )

    columnas_minimas = {"cedula", "nombres_completos", "programa"}
    faltantes = sorted(columnas_minimas - campos_encontrados)
    if faltantes:
        libro.close()
        raise ValueError(
            "La matriz debe incluir las columnas mínimas para importar estudiantes: "
            + ", ".join(
                [
                    "CEDULA",
                    "NOMBRES COMPLETOS",
                    "PROGRAMA",
                ]
            )
            + ". Faltan: " + ", ".join(faltantes)
        )

    creados = 0
    actualizados = 0
    ignoradas = 0
    errores = []
    identificadores_procesados = set()

    for numero_fila, valores_fila in enumerate(
        filas[1:],
        start=2,
    ):
        if not any(
            valor not in (None, "")
            for valor in valores_fila
        ):
            ignoradas += 1
            continue

        datos = {}
        for indice, valor in enumerate(valores_fila):
            encabezado = encabezados_originales[indice]
            campo = resolver_campo(encabezado)
            if campo:
                datos[campo] = valor

        cedula_fila = convertir_cedula(
            obtener(datos, "CEDULA")
        )
        id_banner_fila = convertir_texto(
            obtener(datos, "ID_BANNER")
        )
        identificador_fila = cedula_fila or id_banner_fila

        if not cedula_fila and not id_banner_fila:
            errores.append(
                {
                    "fila": numero_fila,
                    "error": "La cédula está vacía y no existe un ID_BANNER válido para identificar al estudiante.",
                }
            )
            ignoradas += 1
            continue

        if identificador_fila in identificadores_procesados:
            errores.append(
                {
                    "fila": numero_fila,
                    "error": (
                        f"El estudiante con identificador {identificador_fila} "
                        "está repetido dentro del Excel."
                    ),
                }
            )
            continue

        try:
            with transaction.atomic():
                registro, creado = guardar_registro(
                    datos,
                    usuario=usuario,
                )

                identificadores_procesados.add(
                    registro.cedula or registro.id_banner or identificador_fila
                )

                if creado:
                    creados += 1
                else:
                    actualizados += 1

        except (
            ValueError,
            ValidationError,
        ) as error:
            if hasattr(
                error,
                "message_dict",
            ):
                mensaje = "; ".join(
                    (
                        f"{campo}: "
                        f"{', '.join(mensajes)}"
                    )
                    for campo, mensajes
                    in error.message_dict.items()
                )
            else:
                mensaje = str(error)

            errores.append(
                {
                    "fila": numero_fila,
                    "error": mensaje,
                }
            )

        except Exception as error:
            logger.exception(
                "Error inesperado al procesar la fila %s",
                numero_fila,
            )
            errores.append(
                {
                    "fila": numero_fila,
                    "error": "No fue posible procesar esta fila.",
                }
            )

    libro.close()

    return {
        "creados": creados,
        "actualizados": actualizados,
        "ignoradas": ignoradas,
        "errores": errores,
        "total_correctos": creados + actualizados,
    }


def exportar_registro_excel(registro):
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Matriz de titulación"
    hoja.append([encabezado for encabezado, campo in EXPORTAR_CAMPOS])

    fila = []
    for encabezado, campo in EXPORTAR_CAMPOS:
        valor = getattr(registro, campo, "")
        if campo == "fecha_grado" and valor:
            valor = valor.strftime("%d/%m/%Y")
        fila.append(valor if valor is not None else "")

    hoja.append(fila)
    hoja.freeze_panes = "A2"

    for columna in hoja.columns:
        ancho = min(max(len(str(celda.value or "")) for celda in columna) + 2, 40)
        hoja.column_dimensions[columna[0].column_letter].width = ancho

    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    return salida.getvalue()
