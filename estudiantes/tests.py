import io
from zipfile import ZipFile

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook
from openpyxl import load_workbook

from .import_forms import ImportarExcelForm
from .models import HistorialExpediente, Programa, RegistroTitulacion
from .views import registrar_cambios
from .services_excel import convertir_nota, importar_excel
from .services_exportacion import crear_excel_estudiantes


class ImportarExcelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="testuser",
            email="test@example.com",
            password="testpass123",
        )

    def _crear_archivo_excel(self, filas):
        libro = Workbook()
        hoja = libro.active
        hoja.append(filas[0])
        for fila in filas[1:]:
            hoja.append(fila)
        buffer = io.BytesIO()
        libro.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            "estudiantes.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_rejects_non_xlsx_files(self):
        archivo = SimpleUploadedFile(
            "estudiantes.csv",
            b"nombre,cedula",
            content_type="text/csv",
        )

        form = ImportarExcelForm(files={"archivo": archivo})

        self.assertFalse(form.is_valid())
        self.assertIn(".xlsx", form.errors["archivo"][0])

    def test_rejects_notes_outside_supported_range(self):
        with self.assertRaises(ValueError):
            convertir_nota("11")

    def test_import_valid_xlsx_creates_students(self):
        filas = [
            [
                "CEDULA",
                "NOMBRES COMPLETOS",
                "ID_BANNER",
                "CELULAR",
                "CORREO_PERSONAL",
                "CORREO_INSTITUC",
                "SEDE",
                "PROGRAMA",
                "PROGAMA_DESC",
                "ESTADO",
            ],
            [
                "0102030405",
                "Juan Pérez",
                "12345",
                "0999999999",
                "juan@gmail.com",
                "juan@puce.edu.ec",
                "Quito",
                "TQ02",
                "TG Desarrollo de Software",
                "REGISTRADO",
            ],
            [
                "0102030406",
                "María López",
                "12346",
                "0988888888",
                "maria@gmail.com",
                "maria@puce.edu.ec",
                "Quito",
                "TQ01",
                "TG Administración OEPS",
                "EN_PROCESO",
            ],
        ]

        archivo = self._crear_archivo_excel(filas)
        resultado = importar_excel(archivo)

        self.assertEqual(resultado["creados"], 2)
        self.assertEqual(resultado["actualizados"], 0)
        self.assertEqual(RegistroTitulacion.objects.count(), 2)
        self.assertTrue(RegistroTitulacion.objects.filter(cedula="0102030405").exists())
        self.assertTrue(Programa.objects.filter(codigo="TQ02").exists())

    def test_import_updates_existing_student_without_duplicate(self):
        RegistroTitulacion.objects.create(
            cedula="0102030405",
            nombres_completos="Juan Pérez",
            programa="TQ02",
            programa_desc="TG Desarrollo de Software",
            correo_personal="juan@viejo.com",
            correo_instituc="juan@puce.edu.ec",
            estado="REGISTRADO",
        )

        filas = [
            [
                "CEDULA",
                "NOMBRES COMPLETOS",
                "ID_BANNER",
                "SEDE",
                "PROGRAMA",
                "PROGAMA_DESC",
                "ESTADO",
            ],
            [
                "0102030405",
                "Juan Pérez Actualizado",
                "12345",
                "Guayaquil",
                "TQ02",
                "TG Desarrollo de Software",
                "APROBADO",
            ],
        ]

        archivo = self._crear_archivo_excel(filas)
        resultado = importar_excel(archivo)

        self.assertEqual(resultado["creados"], 0)
        self.assertEqual(resultado["actualizados"], 1)
        self.assertEqual(RegistroTitulacion.objects.filter(cedula="0102030405").count(), 1)
        self.assertEqual(
            RegistroTitulacion.objects.get(cedula="0102030405").nombres_completos,
            "Juan Pérez Actualizado",
        )

    def test_import_rejects_empty_xlsx_or_missing_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["CEDULA", "NOMBRES COMPLETOS"])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "vacio.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with self.assertRaises(ValueError):
            importar_excel(archivo)

        empty_file = SimpleUploadedFile(
            "vacío.xlsx",
            b"",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with self.assertRaises(ValueError):
            importar_excel(empty_file)

    def test_import_counts_ignored_rows_and_errors(self):
        filas = [
            [
                "CEDULA",
                "NOMBRES COMPLETOS",
                "PROGRAMA",
                "PROGAMA_DESC",
            ],
            ["", "Ana García", "TQ02", "TG Desarrollo de Software"],
            ["0102030407", "Pedro Paredes", "TQ02", "TG Desarrollo de Software"],
            ["0102030407", "Pedro Paredes Duplicado", "TQ02", "TG Desarrollo de Software"],
            ["0102030408", "", "TQ02", "TG Desarrollo de Software"],
        ]

        archivo = self._crear_archivo_excel(filas)
        resultado = importar_excel(archivo)

        self.assertEqual(resultado["creados"], 1)
        self.assertEqual(resultado["actualizados"], 0)
        self.assertEqual(resultado["ignoradas"], 1)
        self.assertGreater(len(resultado["errores"]), 0)

    def test_importar_excel_get_request_returns_200(self):
        """GET /estudiantes/importar-excel/ debe devolver 200 y mostrar el formulario."""
        self.client.login(username="testuser", password="testpass123")
        response = self.client.get("/estudiantes/importar-excel/")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Importar estudiantes", response.content)
        self.assertIn(b"Procesar matriz", response.content)


class ExportacionTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="exportador",
            password="testpass123",
        )
        self.registros = [
            RegistroTitulacion.objects.create(
                cedula="0102030405",
                id_banner="12345",
                nombres_completos="Juan Pérez",
                programa="TQ02",
                programa_desc="TG Desarrollo de Software",
                estado="APROBADO",
            ),
            RegistroTitulacion.objects.create(
                cedula="0102030406",
                id_banner="12346",
                nombres_completos="María López",
                programa="TQ01",
                programa_desc="TG Administración OEPS",
                estado="EN_PROCESO",
            ),
        ]

    def test_individual_excel_contains_matrix_headers_and_formatting(self):
        self.client.force_login(self.user)
        response = self.client.get(
            f"/estudiantes/{self.registros[0].pk}/exportar-excel/"
        )
        libro = load_workbook(io.BytesIO(response.content))
        hoja = libro.active
        self.assertEqual(response.status_code, 200)
        self.assertEqual(hoja[1][0].value, "ID BANNER")
        self.assertEqual(hoja[2][1].value, "Juan Pérez")
        self.assertEqual(hoja.auto_filter.ref, hoja.dimensions)
        self.assertEqual(HistorialExpediente.objects.count(), 0)

    def test_bulk_formats_are_read_only_and_zip_per_student(self):
        self.client.force_login(self.user)
        base = {
            "estudiante_ids": [str(registro.pk) for registro in self.registros],
        }
        response = self.client.post(
            "/estudiantes/descargar-seleccionados/",
            {**base, "formato": "excel"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            load_workbook(io.BytesIO(response.content)).active.max_row,
            3,
        )

        for formato, extension in (("pdf", ".pdf"), ("word", ".docx")):
            response = self.client.post(
                "/estudiantes/descargar-seleccionados/",
                {**base, "formato": formato},
            )
            with ZipFile(io.BytesIO(response.content)) as archivo_zip:
                nombres = archivo_zip.namelist()
                self.assertEqual(len(nombres), 2)
                self.assertTrue(all(nombre.endswith(extension) for nombre in nombres))
        self.assertEqual(HistorialExpediente.objects.count(), 0)

    def test_lookup_by_banner_cedula_missing_and_login(self):
        self.client.force_login(self.user)
        for parametro, valor in (("id_banner", "12345"), ("cedula", "0102030405")):
            response = self.client.get(f"/estudiantes/buscar/?{parametro}={valor}")
            self.assertEqual(
                response.json()["estudiante"]["nombres_completos"],
                "Juan Pérez",
            )
        self.assertFalse(
            self.client.get(
                "/estudiantes/buscar/?cedula=9999999999"
            ).json()["encontrado"]
        )
        self.client.logout()
        self.assertEqual(
            self.client.get("/estudiantes/buscar/?cedula=0102030405").status_code,
            302,
        )


class AuditoriaTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username="auditor",
            password="testpass123",
            is_staff=True,
        )
        self.user = User.objects.create_user(
            username="operador",
            password="testpass123",
        )
        Programa.objects.get_or_create(
            codigo="TQ02",
            defaults={"descripcion": "TG Desarrollo de Software"},
        )
        self.registro = RegistroTitulacion.objects.create(
            cedula="0102030405",
            nombres_completos="Juan Pérez",
            programa="TQ02",
            programa_desc="TG Desarrollo de Software",
            correo_personal="juan@gmail.com",
            estado="EN_PROCESO",
        )

    def test_manual_creation_generates_audit_entry(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            "/estudiantes/nuevo/",
            {
                "cedula": "0102030406",
                "nombres_completos": "Ana López",
                "programa": "TQ02",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            HistorialExpediente.objects.filter(
                registro__cedula="0102030406",
                accion="CREACION",
                responsable="auditor",
            ).exists()
        )

    def test_creation_and_edit_record_each_changed_field(self):
        anterior = RegistroTitulacion.objects.get(pk=self.registro.pk)
        self.registro.nombres_completos = "Juan Pérez Actualizado"
        self.registro.estado = "APROBADO"
        self.registro.save()

        registrar_cambios(self.registro, anterior, self.staff)

        cambios = HistorialExpediente.objects.filter(registro=self.registro)
        self.assertEqual(cambios.count(), 2)
        self.assertSetEqual(
            set(cambios.values_list("campo", flat=True)),
            {"NOMBRES COMPLETOS", "ESTADO"},
        )
        self.assertTrue(all(cambio.responsable == "auditor" for cambio in cambios))

    def test_import_new_and_existing_records_audit_user_and_fields(self):
        filas = [
            ["CEDULA", "NOMBRES COMPLETOS", "PROGRAMA", "ESTADO"],
            ["0102030406", "Ana López", "TQ02", "REGISTRADO"],
            ["0102030405", "Juan Pérez", "TQ02", "APROBADO"],
        ]
        archivo = ImportarExcelTests()._crear_archivo_excel(filas)
        importar_excel(archivo, usuario=self.staff)

        self.assertTrue(
            HistorialExpediente.objects.filter(
                registro__cedula="0102030406",
                accion="IMPORTACION",
                responsable="auditor",
            ).exists()
        )
        self.assertTrue(
            HistorialExpediente.objects.filter(
                registro=self.registro,
                campo="ESTADO",
                valor_anterior="EN_PROCESO",
                valor_nuevo="APROBADO",
                accion="EDICION",
                responsable="auditor",
            ).exists()
        )

    def test_deletion_keeps_audit_snapshot(self):
        self.client.force_login(self.staff)
        response = self.client.post(f"/estudiantes/{self.registro.pk}/eliminar/")

        self.assertEqual(response.status_code, 302)
        cambio = HistorialExpediente.objects.get(accion="ELIMINACION")
        self.assertIsNone(cambio.registro)
        self.assertEqual(cambio.registro_nombre, "Juan Pérez")
        self.assertEqual(cambio.registro_cedula, "0102030405")

    def test_auditoria_filters_and_individual_history(self):
        anterior = RegistroTitulacion.objects.get(pk=self.registro.pk)
        self.registro.estado = "APROBADO"
        self.registro.save()
        registrar_cambios(self.registro, anterior, self.staff)
        self.client.force_login(self.staff)

        response = self.client.get("/estudiantes/auditoria/?accion=EDICION&cedula=0102030405")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ESTADO")
        self.assertContains(response, "APROBADO")
        response = self.client.get(f"/estudiantes/{self.registro.pk}/historial/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Juan Pérez")

    def test_auditoria_requires_staff(self):
        self.client.force_login(self.user)
        response = self.client.get("/estudiantes/auditoria/")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, "/cuentas/login/?next=/estudiantes/auditoria/")

    def test_revert_creates_new_audit_entry(self):
        anterior = RegistroTitulacion.objects.get(pk=self.registro.pk)
        self.registro.estado = "APROBADO"
        self.registro.save()
        registrar_cambios(self.registro, anterior, self.staff)
        cambio = HistorialExpediente.objects.get(campo="ESTADO")

        self.client.force_login(self.staff)
        response = self.client.post(f"/estudiantes/auditoria/{cambio.pk}/revertir/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            RegistroTitulacion.objects.get(pk=self.registro.pk).estado,
            "EN_PROCESO",
        )
        self.assertTrue(
            HistorialExpediente.objects.filter(
                registro=self.registro,
                accion="REVERSIÓN",
                valor_anterior="APROBADO",
                valor_nuevo="EN_PROCESO",
                responsable="auditor",
            ).exists()
        )


class ImportarExcelViewTests(ImportarExcelTests):

    def test_importar_excel_get_request_contains_form(self):
        """El formulario debe estar presente en GET."""
        self.client.login(username="testuser", password="testpass123")
        response = self.client.get("/estudiantes/importar-excel/")
        self.assertIn(b"multipart/form-data", response.content)

    def test_importar_excel_requires_login(self):
        """Acceso sin autenticación debe redirigir a login."""
        response = self.client.get("/estudiantes/importar-excel/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/cuentas/login/", response.url)

    def test_post_with_valid_excel_creates_students(self):
        """POST con un Excel válido debe crear estudiantes y mostrar resultado."""
        self.client.login(username="testuser", password="testpass123")

        filas = [
            [
                "CEDULA",
                "NOMBRES COMPLETOS",
                "PROGRAMA",
                "PROGAMA_DESC",
            ],
            [
                "0102030409",
                "Carlos López",
                "TQ02",
                "TG Desarrollo de Software",
            ],
        ]

        archivo = self._crear_archivo_excel(filas)
        response = self.client.post(
            "/estudiantes/importar-excel/",
            {"archivo": archivo},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(RegistroTitulacion.objects.filter(cedula="0102030409").exists())

    def test_post_without_file_shows_form(self):
        """POST sin archivo debe mostrar formulario con error."""
        self.client.login(username="testuser", password="testpass123")
        response = self.client.post("/estudiantes/importar-excel/", {})
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Importar estudiantes", response.content)

